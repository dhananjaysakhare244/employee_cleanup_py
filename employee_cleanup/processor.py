from openpyxl import load_workbook

from .excel_utils import find_column, norm


def get_terminated_employees(ws, user_id_name, email_name, terminated_name, terminated_value):
    user_id_col = find_column(ws, user_id_name)
    email_col = find_column(ws, email_name)
    terminated_col = find_column(ws, terminated_name)

    missing = [
        name for name, col in [
            (user_id_name, user_id_col),
            (email_name, email_col),
            (terminated_name, terminated_col),
        ] if not col
    ]
    if missing:
        raise ValueError(
            "Missing employee-file column(s): " + ", ".join(missing)
        )

    ids = set()
    emails = set()
    count = 0

    for row in range(2, ws.max_row + 1):
        if norm(ws.cell(row, terminated_col).value) == norm(terminated_value):
            count += 1
            uid = norm(ws.cell(row, user_id_col).value)
            email = norm(ws.cell(row, email_col).value)

            if uid:
                ids.add(uid)
            if email:
                emails.add(email)

    return ids, emails, count


def process_files(
    employee_file,
    records_file,
    output_file,
    employee_sheet,
    records_sheet,
    employee_user_id,
    employee_email,
    employee_terminated,
    terminated_value,
    records_user_id,
    records_email,
    blank_columns,
):
    employee_wb = load_workbook(employee_file, data_only=False)
    records_wb = load_workbook(records_file, data_only=False)

    if employee_sheet not in employee_wb.sheetnames:
        raise ValueError(f"Employee sheet '{employee_sheet}' was not found.")

    if records_sheet not in records_wb.sheetnames:
        raise ValueError(f"Records sheet '{records_sheet}' was not found.")

    employee_ws = employee_wb[employee_sheet]
    records_ws = records_wb[records_sheet]

    terminated_ids, terminated_emails, terminated_count = (
        get_terminated_employees(
            employee_ws,
            employee_user_id,
            employee_email,
            employee_terminated,
            terminated_value,
        )
    )

    records_user_id_col = find_column(records_ws, records_user_id)
    records_email_col = find_column(records_ws, records_email)

    if not records_user_id_col or not records_email_col:
        missing = []

        if not records_user_id_col:
            missing.append(records_user_id)

        if not records_email_col:
            missing.append(records_email)

        raise ValueError(
            "Missing records-file column(s): " + ", ".join(missing)
        )

    employee_rows = {}

    for row in range(2, records_ws.max_row + 1):
        user_id = norm(
            records_ws.cell(row, records_user_id_col).value
        )

        email = norm(
            records_ws.cell(row, records_email_col).value
        )

        employee_key = None

        if user_id and user_id in terminated_ids:
            employee_key = f"id:{user_id}"
        elif email and email in terminated_emails:
            employee_key = f"email:{email}"

        if employee_key:
            employee_rows.setdefault(
                employee_key,
                []
            ).append(row)

    rows_to_delete = []
    rows_to_retain = []

    for rows in employee_rows.values():
        if rows:
            last_row = rows[-1]

            rows_to_retain.append(last_row)
            rows_to_delete.extend(
                rows[:-1]
            )

    for row in sorted(rows_to_delete, reverse=True):
        records_ws.delete_rows(row, 1)

    for original_row in sorted(rows_to_retain):
        deleted_before = sum(
            1
            for deleted_row in rows_to_delete
            if deleted_row < original_row
        )

        new_row = original_row - deleted_before

        for column in range(1, records_ws.max_column + 1):
            if column != records_user_id_col:
                records_ws.cell(
                    row=new_row,
                    column=column
                ).value = None

    if "Processing Log" in records_wb.sheetnames:
        del records_wb["Processing Log"]

    log = records_wb.create_sheet("Processing Log")

    log.append(["Metric", "Value"])

    matched_employees = len(employee_rows)

    log.append([
        "Terminated employees in employee file",
        terminated_count,
    ])

    log.append([
        "Terminated employees found in records",
        matched_employees,
    ])

    log.append([
        "Records deleted",
        len(rows_to_delete),
    ])

    log.append([
        "Records retained",
        len(rows_to_retain),
    ])

    log.append([
        "Terminated employees not found in records",
        max(
            terminated_count - matched_employees,
            0
        ),
    ])

    log.append([
        "Retained row behavior",
        "User ID kept, all other columns cleared",
    ])

    records_wb.save(output_file)

    return {
        "terminated": terminated_count,
        "matched": matched_employees,
        "deleted": len(rows_to_delete),
        "blanked": len(rows_to_retain),
        "not_found": max(
            terminated_count - matched_employees,
            0
        ),
    }

